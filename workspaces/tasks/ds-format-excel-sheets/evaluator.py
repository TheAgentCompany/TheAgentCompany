import logging
import os
import subprocess
from sotopia.database import AgentProfile

from typing import List

from scoring import Result, Checkpoint
from common import *

import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill

@grader
def grade_checkpoint1():
    try:
        result = subprocess.run(['curl', '-I', 'https://ogma.lti.cs.cmu.edu/apps/files/files/1555?dir=/Documents/Data%20Analysis&openfile=true'], capture_output=True, text=True)
        if result.returncode == 0:
            logging.info("Agent visited the file link on NextCloud.")
            return True
        else:
            logging.warning("Agent did not visit the file link on NextCloud.")
            return False
    except Exception as e:
        logging.error(f"Error occurred while checking the file link on NextCloud: {e}")
        return False

@grader
def grade_checkpoint2():
    try:
        file_url = 'https://ogma.lti.cs.cmu.edu/s/7gGd5AQ5M7WDsL8/download/respondent_counts.xlsx'
        file_path = '/workspace/respondent_counts.xlsx'
        init_file_path = '/utils/respondent_counts.xlsx'
        
        # Download the file from NextCloud
        if not download_nextcloud_content(file_url, file_path):
            logging.warning("Failed to download the Excel file from NextCloud.")
            return False

        df_init = pd.read_excel(init_file_path, sheet_name="unformatted")
        try:
            df = pd.read_excel(file_path, sheet_name="unformatted")
        except:
            logging.warning("Cannot read the 'unformatted' sheet.")
            return False
        
        # Check if the 'unformatted' sheet remains unchanged
        if df_init.equals(df):
            logging.info("The 'unformatted' sheet remains unchanged.")
            return True
        else:
            logging.warning("The 'unformatted' sheet has been changed.")
            return False
    except Exception as e:
        logging.error(f"Error occurred while checking the 'unformatted' sheet: {e}")
        return False

@grader
def grade_checkpoint3():
    try:
        file_path = '/workspace/respondent_counts.xlsx'
        if not os.path.exists(file_path):
            logging.warning("Excel file not found in the data folder.")
            return False
        
        try:
            df_orig = pd.read_excel(file_path, sheet_name="unformatted")
        except:
            logging.warning("Cannot read the 'unformatted' sheet.")
            return False

        try:
            df = pd.read_excel(file_path, sheet_name="formatted")
        except:
            logging.warning("Cannot read the 'formatted' sheet.")
            return False
        
        # The content of the 'formatted' sheet should remain unchanged
        if df_orig.equals(df):
            logging.info("The content of the 'formatted' sheet remains unchanged.")
            return True
        else:
            logging.warning("The content of the 'formatted' sheet has been changed.")
            return False

    except Exception as e:
        logging.error(f"Error occurred while checking the 'formatted' sheet: {e}")
        return False

@grader
def grade_checkpoint4():
    try:
        file_path = '/workspace/respondent_counts.xlsx'
        if not os.path.exists(file_path):
            logging.warning("Excel file not found in the data folder.")
            return False
        
        # Load the workbook
        wb = openpyxl.load_workbook(file_path)
        ws_orig = wb['unformatted']
        ws = wb['formatted']
        
        # Access a specific cell
        cells_to_modify = ["A3", "B3", "B5", "C3", "C4", "D4", "D5", "E5", "F3", "F4", "G4", "G5", "H5", "I3", "I4", "J4", "J5", "K5"]
        
        # The background color of cells in 'cells_to_modify' should be #87CEEB
        # The background color of other cells should remain unchanged
        for col in "ABCDEFGHIJK":
            for row in range(1,22):
                cell_name = f"{col}{row}"
                
                cell = ws[cell_name]
                fill = cell.fill
                
                if cell_name in cells_to_modify:
                    if fill.fill_type == 'solid':
                        color_hex = fill.fgColor.rgb  # Get the RGB color code
                        if color_hex != "FF87CEEB":
                            logging.warning(f"Cell {cell_name} has wrong background color: Expected #87CEEB, obtained {color_hex}.")
                            return False
                    else:
                        logging.warning(f"Cell {cell_name} has wrong background color: Expected #87CEEB, obtained 'None'.")
                        
                    continue
                
                # orig cell
                cell_orig = ws_orig[cell_name]
                fill_orig = cell_orig.fill
                
                if fill.fill_type != fill_orig.fill_type:
                    logging.warning(f"Cell {cell_name} should not change the background color.")
                    return False
                
                if fill_orig.fill_type == 'solid':
                    color_hex_orig = fill_orig.fgColor.rgb
                    color_hex = fill.fgColor.rgb
                    
                    if color_hex != color_hex_orig:
                        logging.warning(f"Cell {cell_name} should not change the background color.")
                        return False
                
                
        logging.info("The required background color changes are applied.")
        return True

    except Exception as e:
        logging.error(f"Error occurred while checking the background color: {e}")
        return False
    
@grader
def grade_checkpoint5():
    try:
        file_path = '/workspace/respondent_counts.xlsx'
        if not os.path.exists(file_path):
            logging.warning("Excel file not found in the data folder.")
            return False
        
        # Load the workbook
        wb = openpyxl.load_workbook(file_path)
        ws = wb['formatted']
        
        for col in "ABCDEFGHIJK":
            for row in range(1,22):
                cell_name = f"{col}{row}"
                
                cell = ws[cell_name]
                if cell.alignment.horizontal != 'center':
                    logging.warning(f"Cell {cell_name} is not horizontally centered")
                    return False
                
                return True
                
        logging.info("All the cells are horizontally centered.")
        return True

    except Exception as e:
        logging.error(f"Error occurred while checking the centering of the cells: {e}")
        return False
    
    

def grade_checkpoints(trajectory="") -> Result:
    checkpoints: List[Checkpoint] = []
    result = Result(checkpoints)

    # Checkpoint 1: Agent visited the file link on NextCloud
    passed_checkpoint1 = grade_checkpoint1()
    checkpoints.append(Checkpoint(1, int(passed_checkpoint1)))

    # Checkpoint 2: The 'unformatted' sheet remains unchanged
    passed_checkpoint2 = grade_checkpoint2()
    checkpoints.append(Checkpoint(1, int(passed_checkpoint2)))

    # Checkpoint 3: The 'formatted' sheet is created and the content is the same as 'unformatted'
    passed_checkpoint3 = grade_checkpoint3()
    checkpoints.append(Checkpoint(1, int(passed_checkpoint3)))

    # Checkpoint 4: The required background colors are applied
    passed_checkpoint4 = passed_checkpoint3 and grade_checkpoint4()
    checkpoints.append(Checkpoint(1, int(passed_checkpoint4)))
    
    # Checkpoint 5: The cells are horizontally centered
    passed_checkpoint5 = passed_checkpoint3 and grade_checkpoint5()
    checkpoints.append(Checkpoint(1, int(passed_checkpoint5)))

    return result



